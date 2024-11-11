import re
import os
import subprocess
from openpyxl import load_workbook
from PySide6.QtCore import QObject, Slot, QSettings, QThread, QDir, Signal

class XLS_Maker(QObject):
    
    F_NAME:str
    F_STATUS: bool = False
    S_NAME:str = "setting.ini"
    
    
    def __init__(self):
        QObject.__init__(self)        

    @Slot(str, result = str)
    def setFileName(self, name) -> str:
        self.F_NAME = name
        self.F_STATUS = False
        return self.getFileInfo()
    
    @Slot(result=bool)
    def getStatus(self):
        return self.F_STATUS
        
    
    def getFileInfo(self):
        wb = load_workbook(self.F_NAME)
        sheets = wb.sheetnames
        
        sett = QSettings(self.S_NAME, QSettings.Format.IniFormat)                
        list_name = sett.value("list/lname")
        first_row = int(sett.value("row/first"))
        
        if list_name not in sheets:
            return f"File - {self.F_NAME}\nFile has not sheet with name - {list_name} - NOT OK\n"
        
        idx = sheets.index(list_name)            
        sh = wb[sheets[idx]]
        
        res = "NOT OK"
        if sh.max_column >= 5 and sh.max_row >= (1 + first_row):
            self.F_STATUS = True
            res = "OK"            
            
        return f"File - {self.F_NAME}\nList {list_name} has {sh.max_column} columns and {sh.max_row} rows with data - {res}\n"
    
    
class MetaChange(QThread):
    S_NAME:str = "setting.ini"
    
    progressChange = Signal(float, name="progressChange")
    finishChange = Signal(str, name = 'finishChange')
    
    def __init__(self, parent = None):
        super().__init__(parent)
    
    @Slot(dict)
    def setParams(self, params:dict):
        self.source = params['source']
        self.target = params['target']
        self.xls = params['xls']
        
    def setMetadataExifToolJPEG(self, ofile, nfile, title, description, keywords):    
        command = [
            'exiftool',
            f'-XMP-dc:title={title}',            
            f'-XMP-dc:description={description}',  
            f'-XMP-dc:subject={keywords}', 
            f'-IPTC:Keywords={keywords}',
            f'-filename={nfile}',   
            ofile
        ]    
        subprocess.run(command, capture_output=True, text=True)
        
    def setMetadataExifToolPNG(self, ofile, nfile, title, description, keywords):
        command = [
            'exiftool',
            f'-XMP-dc:title={title}',
            f'-XMP-dc:description={description}',
            f'-XMP-dc:subject={keywords}', 
            f'-filename={nfile}',   
            ofile
        ]
        subprocess.run(command, capture_output=True, text=True)
        
    def changeMetadataEPS(self, old_file:str,  new_file:str, title:str, desc:str, keywords:list):
    
        with open(old_file, 'rb') as file:
            content = file.read()

        title_pattern = re.compile(rb"(<dc:title>\s*<rdf:Alt>\s*<rdf:li xml:lang=\"x-default\">)(.*?)(</rdf:li>\s*</rdf:Alt>\s*</dc:title>)", re.DOTALL)

        if re.search(title_pattern, content):
            content = re.sub(title_pattern, rb'\1' + title.encode('utf-8') + rb'\3', content)
        
        description_pattern = rb'(<dc:description>.*?<rdf:li xml:lang="x-default">)(.*?)(</rdf:li>.*?</dc:description>)'
        if re.search(description_pattern, content):
            content = re.sub(description_pattern, rb'\1' + desc.encode('utf-8') + rb'\3', content)
        else:
            insert_point = re.search(rb'</dc:title>', content)
            if insert_point:
                content = (content[:insert_point.end()] +
                        b'\n<dc:description>\n<rdf:Alt>\n<rdf:li xml:lang="x-default">' +
                        desc.encode('utf-8') +
                        b'</rdf:li>\n</rdf:Alt>\n</dc:description>' +
                        content[insert_point.end():])

        keywords_pattern = rb'(<dc:subject>.*?<rdf:Bag>)(.*?)(</rdf:Bag>.*?</dc:subject>)'
        new_keywords_xml = ''.join([f'<rdf:li>{kw}</rdf:li>' for kw in keywords]).encode('utf-8')
        if re.search(keywords_pattern, content):
            content = re.sub(keywords_pattern, rb'\1' + new_keywords_xml + rb'\3', content)
        else:
            insert_point = re.search(rb'</dc:description>', content)
            if insert_point:
                content = (content[:insert_point.end()] +
                        b'\n<dc:subject>\n<rdf:Bag>\n' +
                        new_keywords_xml +
                        b'\n</rdf:Bag>\n</dc:subject>' +
                        content[insert_point.end():])


        with open(new_file, "wb") as file:
            file.write(content)
        
    def run(self):     
        self.png: int = 0
        self.jpg: int = 0
        self.eps: int = 0
           
        source_dir = QDir(QDir.toNativeSeparators(self.source))
        source_list = source_dir.entryList([QDir.Filter.Files, "*.png", "*.jpg", "*.eps"])
        
        wb = load_workbook(self.xls)
        sheets = wb.sheetnames
        
        
        sett = QSettings(self.S_NAME, QSettings.Format.IniFormat)                
        list_name = sett.value("list/lname")
        start_row = int(sett.value("row/first"))
        
        
        col_old_file = int(sett.value("col/old_file"))
        col_new_file = int(sett.value("col/new_file"))
        col_title = int(sett.value("col/title"))
        col_desc = int(sett.value("col/desc"))
        col_keys = int(sett.value("col/keys"))
        
        columns = [col_old_file, col_new_file, col_title, col_desc, col_keys]
        
        key_separator = sett.value("col/separator")
        
        if list_name not in sheets:
            print("Error list")
            return
        idx = sheets.index(list_name)            
        sh = wb[sheets[idx]]
        end_row = sh.max_row
        
        r:int = 0
        for row in sh.iter_rows(min_row=start_row, max_row=end_row, values_only=True):            
            
            e_list = [".eps", ".png", ".jpg"]
            
            for col in columns:
                value = row[col-1]                
                if col == col_old_file:
                    o_name  = value                    
                if col == col_new_file:
                    n_name = value
                if col == col_title:
                    title = value
                if col == col_desc:
                    desc = value
                if col == col_keys:
                    keys = value
            
            for e in e_list:
                fn = f"{o_name}{e}"
                if fn in source_list:
                    
                    oldFile = QDir.toNativeSeparators(f"{self.source}/{o_name}{e}")
                    newFile = QDir.toNativeSeparators(f"{self.target}/{n_name}{e}")
                    
                    if e == ".eps":
                        keywords = keys.split(key_separator)
                        self.changeMetadataEPS(oldFile, newFile, title, desc, keywords)
                        self.eps += 1
                    if e == ".png":
                        self.setMetadataExifToolPNG(oldFile, newFile, title, desc, keys)
                        self.png += 1
                    if e == ".jpg":
                        self.setMetadataExifToolJPEG(oldFile, newFile, title, desc, keys)
                        self.jpg += 1
            r += 1
            self.progressChange.emit(r / (sh.max_row - start_row))
        
        res = f"\nChanged metadata in {self.png + self.jpg + self.eps} files: PNG: {self.png} | JPG: {self.jpg} | EPS: {self.eps}"
        self.finishChange.emit(res)